import csv
from django import forms
from django.contrib import admin
from django.contrib.auth.forms import UserChangeForm
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _

# =========================================================
# UNFOLD IMPORTS (Strictly Unfold - No Django Overrides)
# =========================================================
from unfold.admin import ModelAdmin, TabularInline
from unfold.decorators import display, action

# =========================================================
# INTERNAL IMPORTS
# =========================================================
from operations.mixins import SecurityAuditMixin
from profiles.models import CompanyMember
from .forms import CustomUserAdminCreationForm
from .models import (
    ApplicationRequest, City, CommunityContributor, Country,
    CurrentStatus, CustomUser, FieldOfInterest, IDSequence,
    Institution, StaffUser, UniversalContactMethod, UniversalSocialLink
)


# =========================================================
# HELPER: DYNAMIC PROFILE RESOLVER
# =========================================================
def get_user_profile(user):
    """
    Safely retrieves the Unified Portfolio.
    Replaces all legacy role-based profile checks.
    """
    try:
        if hasattr(user, 'portfolio'):
            return user.portfolio
    except ObjectDoesNotExist:
        pass
    return None


# =========================================================
# 1. CUSTOM FORMS
# =========================================================

class CustomUserChangeForm(forms.ModelForm):
    password = forms.CharField(
        widget=forms.PasswordInput(),
        required=False,
        help_text=_("Leave blank to keep current password. Type a new one here to reset it.")
    )

    admin_rating = forms.IntegerField(
        min_value=0, max_value=5, required=False,
        label=_("Profile Admin Rating"),
        help_text=_("Curate the user's rating (0-5). Automatically syncs to their Unified Portfolio.")
    )

    is_rating_locked = forms.BooleanField(
        required=False,
        label=_("Admin Rating Lock"),
        help_text=_("Check this to stop the AI Oracle from auto-updating this rating.")
    )

    profile_verified = forms.BooleanField(
        required=False,
        label=_("Profile Verified"),
        help_text=_("Check this to mark the user's profile as verified on the public portfolio page.")
    )

    class Meta:
        model = CustomUser
        fields = '__all__'
        field_classes = {'phone_number': forms.CharField}

    def clean_email(self):
        email = self.cleaned_data.get('email')
        return email if email else None

    def clean_phone_number(self):
        phone = self.cleaned_data.get('phone_number')
        return phone if phone else None

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            profile = get_user_profile(self.instance)
            if profile:
                self.fields['admin_rating'].initial = profile.admin_rating
                self.fields['is_rating_locked'].initial = getattr(profile, 'is_rating_locked', False)
                self.fields['profile_verified'].initial = getattr(profile, 'profile_verified', False)
            else:
                self.fields['admin_rating'].disabled = True
                self.fields['is_rating_locked'].disabled = True
                self.fields['profile_verified'].disabled = True
                self.fields['admin_rating'].help_text = _(
                    "User does not have an active Unified Portfolio yet. Cannot set rating.")
                self.fields['is_rating_locked'].help_text = _("Portfolio required to lock rating.")
                self.fields['profile_verified'].help_text = _("Portfolio required to verify profile.")

    def save(self, commit=True):
        user = super().save(commit=False)
        if self.cleaned_data.get("password"):
            user.set_password(self.cleaned_data["password"])
        if commit:
            user.save()
            if hasattr(self, 'save_m2m'):
                self.save_m2m()
        return user


class StaffUserForm(forms.ModelForm):
    password = forms.CharField(
        widget=forms.PasswordInput(),
        required=False,
        help_text=_("Leave blank to keep current password.")
    )

    class Meta:
        model = StaffUser
        fields = ('full_name', 'phone_number', 'password', 'is_active', 'is_staff', 'is_superuser', 'groups')

    def save(self, commit=True):
        user = super().save(commit=False)
        if self.cleaned_data.get("password"):
            user.set_password(self.cleaned_data["password"])
        if commit:
            user.save()
            self.save_m2m()
        return user


# =========================================================
# 2. INLINES
# =========================================================

class CompanyMemberInline(TabularInline):
    model = CompanyMember
    extra = 0
    fields = ['company', 'role', 'job_title', 'is_active']
    readonly_fields = ['company']
    can_delete = False
    show_change_link = True
    tab = True


class SocialLinkInline(TabularInline):
    model = UniversalSocialLink
    extra = 1
    can_delete = True
    tab = True
    fields = ('platform_name', 'url', 'order')


class ContactMethodInline(TabularInline):
    model = UniversalContactMethod
    extra = 1
    can_delete = True
    tab = True
    fields = ('type', 'value')


class ApplicationRequestInline(TabularInline):
    model = ApplicationRequest
    extra = 0
    tab = True

    readonly_fields = ['submission_data', 'created_at', 'safe_download_cv', 'cv_file']
    fields = ['role_type', 'status', 'safe_download_cv', 'admin_notes', 'created_at']
    can_delete = False

    @display(description=_("CV Action"))
    def safe_download_cv(self, obj):
        if obj.cv_file:
            return format_html(
                '<a href="{}" target="_blank" style="display: inline-block; padding: 6px 14px; background-color: #0A66C2; color: white; font-weight: bold; border-radius: 6px; text-decoration: none; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;">⬇️ Open CV</a>',
                obj.cv_file.url
            )
        return mark_safe('<span style="color: #9ca3af; font-size: 11px; font-weight: bold;">NO CV</span>')


class CityInline(TabularInline):
    model = City
    extra = 1
    fields = ['name', 'slug', 'is_verified']
    prepopulated_fields = {'slug': ('name',)}
    show_change_link = True


# =========================================================
# 3. MAIN USER ADMIN
# =========================================================

@admin.register(CustomUser)
class CustomUserAdmin(SecurityAuditMixin, ModelAdmin):
    add_form = CustomUserAdminCreationForm
    form = CustomUserChangeForm
    ordering = ('-date_joined',)

    # =========================================================
    # BOTH TOP-LEVEL ACTION BUTTONS
    # =========================================================
    actions_list = ["export_users_html", "export_users_excel"]

    # --- BUTTON 1: THE HTML CRM REPORT ---
    @action(description=_("🌐 Open HTML Report"), url_path="export-users-html")
    def export_users_html(self, request):
        html_content = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>CoreLink | Executive Talent Report</title>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap" rel="stylesheet">
            <style>
                body { font-family: 'Inter', sans-serif; background-color: #F8FAFC; color: #0F172A; padding: 40px; margin: 0; }
                .header { text-align: center; margin-bottom: 30px; }
                .header h1 { color: #0A66C2; font-weight: 800; font-size: 28px; margin: 0; letter-spacing: -0.5px; }
                .header p { color: #64748B; font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 1px; margin-top: 5px; }

                .table-wrapper { background: #FFFFFF; border-radius: 12px; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); overflow: hidden; border: 1px solid #E2E8F0; }
                table { width: 100%; border-collapse: collapse; text-align: left; }
                th { background-color: #0A66C2; color: #FFFFFF; padding: 16px; font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; }
                td { padding: 14px 16px; border-bottom: 1px solid #E2E8F0; font-size: 14px; color: #334155; vertical-align: middle; }
                tr:last-child td { border-bottom: none; }
                tr:nth-child(even) { background-color: #F8FAFC; }
                tr:hover { background-color: #F0F6FF; }

                .avatar-img { width: 36px; height: 36px; border-radius: 50%; object-fit: cover; border: 1px solid #CBD5E1; box-shadow: 0 2px 4px rgba(0,0,0,0.05); display: block; }

                .btn { display: inline-block; padding: 6px 12px; border-radius: 6px; font-size: 11px; font-weight: bold; text-decoration: none; text-transform: uppercase; transition: all 0.2s; }
                .btn-profile { background-color: #EBF4FD; color: #0A66C2; border: 1px solid #BFDBFE; }
                .btn-profile:hover { background-color: #0A66C2; color: #FFFFFF; }
                .btn-tg { background-color: #F0FDF4; color: #16A34A; border: 1px solid #BBF7D0; }
                .btn-tg:hover { background-color: #16A34A; color: #FFFFFF; }
                .btn-disabled { background-color: #F1F5F9; color: #94A3B8; cursor: not-allowed; }

                .badge-yes { background-color: #DCFCE7; color: #16A34A; padding: 4px 8px; border-radius: 4px; font-size: 11px; font-weight: bold; border: 1px solid #BBF7D0; }
                .badge-no { background-color: #FEF2F2; color: #DC2626; padding: 4px 8px; border-radius: 4px; font-size: 11px; font-weight: bold; border: 1px solid #FECACA; }

                @media print {
                    body { padding: 0; background-color: white; }
                    .table-wrapper { box-shadow: none; border: none; }
                    th { background-color: #F1F5F9 !important; color: #0F172A !important; -webkit-print-color-adjust: exact; }
                    .btn { border: 1px solid #CBD5E1; color: #0F172A; }
                }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>CORELINK ETHIOPIA</h1>
                <p>Executive Talent Database Report</p>
            </div>
            <div class="table-wrapper">
                <table>
                    <thead>
                        <tr>
                            <th>Avatar</th>
                            <th>Status</th>
                            <th>Name</th>
                            <th>Title / Expertise</th>
                            <th>Phone</th>
                            <th>Email</th>
                            <th>Portfolio</th>
                            <th>Direct Contact</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        queryset = CustomUser.objects.all().select_related('portfolio').prefetch_related(
            'portfolio__headlines').order_by('is_contacted', '-date_joined')

        for user in queryset:
            title = "No Title Provided"
            profile = get_user_profile(user)
            if profile:
                primary_headline = profile.headlines.filter(is_primary=True).first()
                if primary_headline:
                    title = primary_headline.title
                else:
                    first_headline = profile.headlines.first()
                    if first_headline:
                        title = first_headline.title

            try:
                profile_path = user.get_absolute_url()
                full_url = request.build_absolute_uri(profile_path)
            except Exception:
                full_url = f"https://corelink.et/profile/{user.corelink_id or user.id}/"

            email = user.email if user.email else "<span style='color:#94A3B8'>N/A</span>"
            phone = user.phone_number if user.phone_number else "<span style='color:#94A3B8'>N/A</span>"

            profile_btn = f'<a href="{full_url}" target="_blank" class="btn btn-profile">View Profile</a>'

            if user.telegram_handle:
                clean_tg = user.telegram_handle.replace("@", "").strip()
                telegram_btn = f'<a href="https://t.me/{clean_tg}" target="_blank" class="btn btn-tg">Message @{clean_tg}</a>'
            else:
                telegram_btn = f'<span class="btn btn-disabled">No Telegram</span>'

            contact_badge = '<span class="badge-yes">✅ Contacted</span>' if user.is_contacted else '<span class="badge-no">⏳ Pending</span>'

            try:
                if user.avatar and hasattr(user.avatar, 'url'):
                    avatar_src = request.build_absolute_uri(user.avatar.url)
                else:
                    avatar_src = f"https://ui-avatars.com/api/?name={user.display_name}&background=EBF4FF&color=7F9CF5&bold=true"
            except Exception:
                avatar_src = f"https://ui-avatars.com/api/?name={user.display_name}&background=EBF4FF&color=7F9CF5&bold=true"

            avatar_html = f'<img src="{avatar_src}" class="avatar-img" alt="{user.display_name}">'

            html_content += f"""
                        <tr>
                            <td style="width: 50px;">{avatar_html}</td>
                            <td>{contact_badge}</td>
                            <td><strong>{user.display_name}</strong></td>
                            <td>{title}</td>
                            <td>{phone}</td>
                            <td>{email}</td>
                            <td>{profile_btn}</td>
                            <td>{telegram_btn}</td>
                        </tr>
            """

        html_content += """
                    </tbody>
                </table>
            </div>
        </body>
        </html>
        """
        return HttpResponse(html_content, content_type='text/html')

    # --- BUTTON 2: THE EXCEL/PDF DOWNLOAD ---
    @action(description=_("📊 Download Excel / PDF"), url_path="export-users-excel")
    def export_users_excel(self, request):
        """
        Generates a true .XLSX Excel file engineered for PDF conversion.
        Forces columns to fit on one landscape page without cutting off.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="CoreLink_Talent_Database.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Talent Pool'

        header_fill = PatternFill(start_color="0A66C2", end_color="0A66C2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
        )
        # CRITICAL FIX: wrap_text=True stops URLs from pushing everything off the page
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        headers = ['Status', 'Full Name', 'Profile Link', 'Professional Title', 'Phone Number', 'Telegram Handle']
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = cell_alignment

        queryset = CustomUser.objects.all().select_related('portfolio').prefetch_related(
            'portfolio__headlines').order_by('is_contacted', '-date_joined')

        for user in queryset:
            title = "No Title Provided"
            profile = get_user_profile(user)
            if profile:
                primary_headline = profile.headlines.filter(is_primary=True).first()
                if primary_headline:
                    title = primary_headline.title
                else:
                    first_headline = profile.headlines.first()
                    if first_headline:
                        title = first_headline.title

            try:
                profile_path = user.get_absolute_url()
                full_url = request.build_absolute_uri(profile_path)
            except Exception:
                full_url = f"https://corelink.et/profile/{user.corelink_id or user.id}/"

            phone = str(user.phone_number) if user.phone_number else "N/A"
            contacted = "Yes" if user.is_contacted else "No"

            if user.telegram_handle:
                clean_tg = user.telegram_handle.replace("@", "").strip()
                telegram = f"https://t.me/{clean_tg}"
            else:
                telegram = "N/A"

            row = [contacted, user.display_name, full_url, title, phone, telegram]
            worksheet.append(row)

            current_row = worksheet.max_row
            for col_idx in range(1, 7):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = cell_alignment

                if col_idx == 3:
                    cell.hyperlink = full_url
                    cell.style = "Hyperlink"
                    cell.border = thin_border
                    cell.alignment = cell_alignment
                elif col_idx == 6 and telegram != "N/A":
                    cell.hyperlink = telegram
                    cell.style = "Hyperlink"
                    cell.border = thin_border
                    cell.alignment = cell_alignment

        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # CRITICAL FIX: Hard cap width at 35 so it never stretches off the PDF page
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 35)

        # PDF PAGE SETUP (The Ironclad Lock)
        worksheet.page_margins.left = 0.2
        worksheet.page_margins.right = 0.2
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        workbook.save(response)
        return response

    # ---------------------------------------------------------
    # UI LIST DISPLAY
    # ---------------------------------------------------------
    list_display = [
        'display_header',
        'contact_details',
        'role_and_rating',
        'is_verified',
        'profile_verified_status',
        'is_nexus_visible',
        'is_selected',
        'is_top_10',
        'is_pinned_in_right_now',
        'is_hero_avatar_selected',
        'is_home_profile_selected',
        'is_active',
    ]

    list_editable = [
        'is_verified',
        'is_nexus_visible',
        'is_selected',
        'is_top_10',
        'is_pinned_in_right_now',
        'is_hero_avatar_selected',
        'is_home_profile_selected',
        'is_active'
    ]

    list_filter = [
        'role',
        'is_hero_avatar_selected',
        'is_home_profile_selected',
        'is_selected',
        'is_top_10',
        'is_pinned_in_right_now',
        'is_nexus_visible',
        'is_verified',
        'is_active',
        'date_joined'
    ]

    search_fields = ['phone_number', 'email', 'full_name', 'telegram_handle', 'corelink_id']
    inlines = [CompanyMemberInline, ContactMethodInline, SocialLinkInline, ApplicationRequestInline]

    fieldsets = (
        (_("Identity & Role"), {
            "fields": (
                'full_name', 'telegram_handle', 'role', 'corelink_id',
                'current_location', 'avatar', 'cover_image', 'is_verified',
                ('admin_rating', 'is_rating_locked', 'profile_verified')
            ),
            "classes": ["tab"]
        }),
        (_("📞 CRM / Contact Status"), {
            "fields": (
                'is_contacted',
            ),
            "description": "Toggle this ON once you or your team have reached out to this user.",
            "classes": ["tab"]
        }),
        (_("🏠 Home Page Curation"), {
            "fields": (
                'is_hero_avatar_selected',
                'is_home_profile_selected',
            ),
            "description": "Control exactly who appears on the public landing page.",
            "classes": ["collapse"]
        }),
        (_("🛡️ Feed Control & Moderation"), {
            "fields": (
                'is_nexus_visible',
                'is_selected',
                'is_top_10',
                'is_pinned_in_right_now',
                'is_banned_from_right_now'
            ),
            "description": "Toggle visibility across the Nexus and Right Now feeds.",
            "classes": ["collapse"]
        }),
        (_("Security & Access"), {
            "fields": ('phone_number', 'email', 'is_email_verified', 'password', 'is_active', 'is_staff',
                       'is_superuser'),
            "classes": ["tab"]
        }),
        (_("Permissions & Groups"), {
            "fields": ('groups', 'user_permissions'),
            "classes": ["tab"]
        }),
        (_("Timestamps"), {
            "fields": ('date_joined', 'last_login'),
            "classes": ["collapse"]
        }),
    )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('portfolio')

    def save_model(self, request, obj, form, change):
        if getattr(obj, 'email', None) == "":
            obj.email = None
        if getattr(obj, 'phone_number', None) == "":
            obj.phone_number = None

        super().save_model(request, obj, form, change)

        profile = get_user_profile(obj)
        if profile:
            rating = form.cleaned_data.get('admin_rating')
            is_locked = form.cleaned_data.get('is_rating_locked')
            # Check both field names (form field and list_editable method name)
            is_verified = form.cleaned_data.get('profile_verified') or form.cleaned_data.get('profile_verified_status')
            update_fields = []

            if rating is not None and getattr(profile, 'admin_rating', None) != rating:
                profile.admin_rating = rating
                update_fields.append('admin_rating')

            if is_locked is not None and getattr(profile, 'is_rating_locked', None) != is_locked:
                profile.is_rating_locked = is_locked
                update_fields.append('is_rating_locked')

            if is_verified is not None and getattr(profile, 'profile_verified', None) != is_verified:
                profile.profile_verified = is_verified
                update_fields.append('profile_verified')

            if update_fields:
                profile.save(update_fields=update_fields)

    STICKY_HEADER_CSS = mark_safe(
        str(_("User Identity")) +
        '<style>'
        '#result_list thead th { position: sticky !important; top: 0 !important; z-index: 40 !important; outline: 1px solid rgba(128,128,128,0.1); } '
        '.dark #result_list thead th { background-color: #111827 !important; } '
        '#result_list thead th { background-color: #f9fafb !important; }'
        '</style>'
    )

    @display(description=STICKY_HEADER_CSS)
    def display_header(self, obj):
        image_url = obj.avatar.url if obj.avatar else f"https://ui-avatars.com/api/?name={obj.full_name}&background=EBF4FF&color=7F9CF5"

        company_html = ""
        memberships = obj.company_memberships.filter(is_active=True).select_related('company')
        if memberships.exists():
            companies = ", ".join([m.company.name for m in memberships])
            company_html = f'<div style="font-size: 10px; font-weight: bold; color: #b8860b; margin-top: 2px;">🏢 {companies}</div>'
        elif str(getattr(obj, 'role', '')).upper() == 'FOUNDER':
            company_html = '<div style="font-size: 10px; font-weight: bold; color: #854d0e; margin-top: 2px;">👑 Independent Founder</div>'

        try:
            profile_url = obj.get_absolute_url()
        except Exception:
            profile_url = f"/profile/{obj.corelink_id or obj.id}/"

        link_html = f'''
            <div style="margin-top: 8px;">
                <a href="{profile_url}" target="_blank" onclick="event.stopPropagation();" 
                   style="display: inline-flex; align-items: center; gap: 3px; font-size: 10px; font-weight: 600; color: #4338ca; text-decoration: none; text-transform: uppercase; letter-spacing: 0.5px;">
                    View Profile 
                    <svg width="10" height="10" fill="none" stroke="currentColor" stroke-width="2.5" viewBox="0 0 24 24" stroke-linecap="round" stroke-linejoin="round">
                        <path d="M7 17l9.2-9.2M17 17V7H7"/>
                    </svg>
                </a>
            </div>
        '''

        return format_html(
            '''
            <div style="display: flex; align-items: flex-start; gap: 10px; min-width: 180px; max-width: 250px; white-space: normal;">
                <img src="{}" style="width: 36px; height: 36px; border-radius: 50%; object-fit: cover; border: 1px solid #eee; flex-shrink: 0;" />
                <div style="display: flex; flex-direction: column; overflow: hidden;">
                    <span style="font-weight: 600; font-size: 13px; line-height: 1.2;">{}</span>
                    <span style="font-size: 11px; color: #6b7280; margin-top: 2px;">{}</span>
                    {}
                    {}
                </div>
            </div>
            ''',
            image_url, obj.full_name, obj.telegram_handle or "",
            mark_safe(company_html), mark_safe(link_html)
        )

    @display(description=_("Contact Info"))
    def contact_details(self, obj):
        email_status = '<span style="color: #16a34a;" title="Verified Email">✅</span>' if obj.is_email_verified else '<span style="color: #d97706;" title="Unverified Email">⏳</span>'
        phone = obj.phone_number or '<span style="color: #9ca3af;">No Phone</span>'
        email = obj.email or '<span style="color: #9ca3af;">No Email</span>'

        return format_html(
            '<div style="font-size: 11px; line-height: 1.5; min-width: 140px; max-width: 200px; white-space: normal; word-wrap: break-word;">'
            '<strong>📞</strong> {}<br>'
            '<div style="margin-top: 4px;"><strong>📧</strong> {} {}</div>'
            '</div>',
            mark_safe(phone), mark_safe(email), mark_safe(email_status)
        )

    @display(description=_("Rating & CV"))
    def role_and_rating(self, obj):
        profile = get_user_profile(obj)
        rating_html = ""
        cv_html = ""

        if profile:
            rating = profile.admin_rating
            is_locked = getattr(profile, 'is_rating_locked', False)
            stars = "⭐" * rating + "☆" * (5 - rating)
            lock_icon = ' <span title="Rating Locked">🔒</span>' if is_locked else ''
            rating_html = f'<div style="color: #ca8a04; font-size: 11px; margin-top: 6px; white-space: nowrap;" title="Rating: {rating}/5">{stars}{lock_icon}</div>'

            # CV indicator
            if profile.cv_file:
                cv_html = '<div style="color: #16a34a; font-size: 10px; margin-top: 4px; font-weight: bold;">📄 CV Uploaded</div>'
            else:
                cv_html = '<div style="color: #9ca3af; font-size: 10px; margin-top: 4px;">No CV</div>'
        else:
            rating_html = '<div style="color: #9ca3af; font-size: 10px; margin-top: 6px;">No Profile</div>'
            cv_html = '<div style="color: #9ca3af; font-size: 10px; margin-top: 4px;">No CV</div>'

        return format_html('<div style="min-width: 90px;">{}{}</div>', mark_safe(rating_html), mark_safe(cv_html))

    @display(description=_("Profile Verified"), boolean=True)
    def profile_verified_status(self, obj):
        profile = get_user_profile(obj)
        if profile:
            return profile.profile_verified
        return False

    def get_readonly_fields(self, request, obj=None):
        if not request.user.is_superuser:
            return [
                'full_name', 'phone_number', 'telegram_handle', 'role', 'corelink_id',
                'current_location', 'avatar', 'cover_image', 'is_active', 'is_staff', 'is_superuser', 'groups',
                'user_permissions', 'date_joined', 'last_login', 'password'
            ]
        return self.readonly_fields

    def get_fieldsets(self, request, obj=None):
        fieldsets = super().get_fieldsets(request, obj)
        if not request.user.is_superuser:
            return [fs for fs in fieldsets if fs[0] == _("Identity & Role")]
        return fieldsets

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def get_inlines(self, request, obj=None):
        if not request.user.is_superuser:
            return [CompanyMemberInline]
        return self.inlines


# =========================================================
# 4. SECONDARY ADMINS
# =========================================================

@admin.register(ApplicationRequest)
class ApplicationRequestAdmin(SecurityAuditMixin, ModelAdmin):
    ordering = ('-created_at',)

    list_display = ['user', 'role_type', 'status_badge', 'download_cv_btn', 'created_at']
    list_filter = ['status', 'role_type', 'created_at']
    search_fields = ['user__full_name', 'user__phone_number']

    readonly_fields = ['submission_data', 'created_at', 'download_cv_btn', 'cv_file']

    fieldsets = (
        (_("🚨 Quick Actions"), {
            "fields": ('download_cv_btn', 'status'),
        }),
        (_("Application Details"), {
            "fields": ('user', 'role_type', 'admin_notes', 'submission_data', 'created_at')
        }),
    )

    @display(description=_("Status"))
    def status_badge(self, obj):
        colors = {"PENDING": ("#ca8a04", "#fef08a"), "APPROVED": ("#16a34a", "#dcfce7"),
                  "REJECTED": ("#dc2626", "#fee2e2")}
        text_color, bg_color = colors.get(obj.status, ("#4b5563", "#f3f4f6"))
        return format_html(
            '<span style="background: {}; color: {}; padding: 4px 8px; border-radius: 6px; font-size: 11px; font-weight: bold;">{}</span>',
            bg_color, text_color, obj.get_status_display()
        )

    @display(description=_("CV Document"))
    def download_cv_btn(self, obj):
        if obj.cv_file:
            return format_html(
                '<a href="{}" target="_blank" style="display: inline-block; padding: 6px 14px; background-color: #0A66C2; color: white; border-radius: 8px; font-weight: 800; text-decoration: none; font-size: 13px; text-transform: uppercase; letter-spacing: 1px; box-shadow: 0 4px 6px -1px rgba(10, 102, 194, 0.2);">'
                '<svg width="18" height="18" fill="none" stroke="currentColor" viewBox="0 0 24 24" stroke-width="2.5"><path stroke-linecap="round" stroke-linejoin="round" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"></path></svg>'
                'View & Download CV'
                '</a>',
                obj.cv_file.url
            )
        return mark_safe(
            '<span style="color: #ef4444; font-weight: bold; padding: 10px; background: #fef2f2; border-radius: 6px; display: inline-block;">No CV Attached to this application.</span>')


@admin.register(CommunityContributor)
class CommunityContributorAdmin(SecurityAuditMixin, ModelAdmin):
    ordering = ('-created_at',)
    list_display = ['full_name', 'telegram_username', 'contribution_area', 'contact_status_badge', 'created_at']
    list_filter = ['is_contacted', 'created_at']
    search_fields = ['full_name', 'telegram_username', 'contribution_area']

    @display(description=_("Contact Status"))
    def contact_status_badge(self, obj):
        if obj.is_contacted:
            return format_html(
                '<span style="background: #dcfce7; color: #16a34a; padding: 4px 8px; border-radius: 6px; font-size: 11px; font-weight: bold;">Contacted</span>')
        return format_html(
            '<span style="background: #fee2e2; color: #dc2626; padding: 4px 8px; border-radius: 6px; font-size: 11px; font-weight: bold;">Waiting</span>')


@admin.register(IDSequence)
class IDSequenceAdmin(ModelAdmin):
    list_display = ['prefix', 'year', 'last_number']


@admin.register(StaffUser)
class StaffUserAdmin(SecurityAuditMixin, ModelAdmin):
    form = StaffUserForm
    list_display = ['display_header', 'phone_number', 'is_active', 'is_superuser']

    def get_queryset(self, request):
        return super().get_queryset(request).filter(is_staff=True)

    @display(description=_("Staff Member"))
    def display_header(self, obj):
        image_url = obj.avatar.url if obj.avatar else f"https://ui-avatars.com/api/?name={obj.full_name}"
        return format_html(
            '<div style="display: flex; align-items: center; gap: 10px;">'
            '<img src="{}" style="width: 32px; height: 32px; border-radius: 50%; object-fit: cover;" /><b>{}</b></div>',
            image_url, obj.full_name
        )


# =========================================================
# 5. LOCATION ADMINS
# =========================================================

@admin.register(Country)
class CountryAdmin(ModelAdmin):
    list_display = ['name', 'slug', 'is_verified', 'city_count', 'created_at']
    list_filter = ['is_verified']
    search_fields = ['name']
    prepopulated_fields = {'slug': ('name',)}
    inlines = [CityInline]

    @display(description=_("Cities"))
    def city_count(self, obj):
        return obj.city_set.count()


@admin.register(City)
class CityAdmin(ModelAdmin):
    list_display = ['name', 'Country', 'slug', 'is_verified', 'created_at']
    list_filter = ['is_verified', 'Country']
    search_fields = ['name', 'Country__name']
    prepopulated_fields = {'slug': ('name',)}
    autocomplete_fields = ['Country']


# =========================================================
# 6. INSTITUTION & FIELD OF INTEREST ADMINS
# =========================================================

@admin.register(Institution)
class InstitutionAdmin(ModelAdmin):
    list_display = ['name', 'City', 'slug', 'is_verified', 'created_at']
    list_filter = ['is_verified', 'City']
    search_fields = ['name', 'City__name']
    prepopulated_fields = {'slug': ('name',)}


@admin.register(FieldOfInterest)
class FieldOfInterestAdmin(ModelAdmin):
    list_display = ['id', 'name']
    search_fields = ['name']


@admin.register(CurrentStatus)
class CurrentStatusAdmin(ModelAdmin):
    list_display = ['id', 'name']
    search_fields = ['name']